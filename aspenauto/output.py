import pyexcel
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill
from .objectcollection import ObjectCollection

class Output(object):
    
    def __init__(self):
        
        return

    
    def Print_Mass(self, material_streams, work_book):
        
        sheet_name = 'Mass balances v2' 
        
        
        wb = load_workbook(work_book)
    
        # if the sheet doesn't exist, create a new sheet
        try:
            sheet = wb[sheet_name]
            wb.remove(sheet)
            sheet = wb.create_sheet(sheet_name)
        except:
            sheet = wb.create_sheet(sheet_name)

        count = 0
        feed_count, feed_mass = 0, 0
        product_count,product_mass = 0, 0
        waste_count, waste_mass = 0, 0

        
        rw = 19

        for name, obj in material_streams.items():
            comp_count = 0

            if obj.type == 'Feed':
                row_start = rw + count
                for comp, frac in obj.massfrac.items():
                
                    if frac >= 0.0001:
                        
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        feed_count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow * 8000 / 1000 / 1000
                feed_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure


            if obj.type == 'Product':
                row_start = rw + count + 3
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:

                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        count += 1
                        product_count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow * 8000 / 1000 / 1000
                product_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure


            if obj.type == 'Waste':
                row_start = rw + count + 6
                for comp, frac in obj.massfrac.items():
                    if frac >= 0.0001:
                        sheet.cell(row=row_start + comp_count , column=4).value = comp
                        sheet.cell(row=row_start + comp_count , column=5).value = frac * 100
                        comp_count += 1
                        waste_count += 1
                        count += 1

                sheet.merge_cells(start_row= row_start, start_column=2, end_row=row_start+comp_count-1, end_column=2)
                sheet.cell(row= row_start , column=2).value = name
                sheet.merge_cells(start_row= row_start, start_column=3, end_row=row_start+comp_count-1, end_column=3)
                massflow = obj.massflow * 8000 / 1000 / 1000
                waste_mass += massflow
                sheet.cell(row= row_start , column=3).value = massflow
                sheet.merge_cells(start_row= row_start, start_column=6, end_row=row_start+comp_count-1, end_column=6)
                sheet.cell(row= row_start , column=6).value = obj.temperature
                sheet.merge_cells(start_row= row_start, start_column=7, end_row=row_start+comp_count-1, end_column=7)
                sheet.cell(row= row_start , column=7).value = obj.pressure
        
        # Print feed headers
        headers = ['Stream','Flowrate ktonne/y','Component','Concentration (wt%)','Temperature ( C )','Pressure (bar)']
        feed_start = rw
        col_count = 2
        self.fill_cell_bold(sheet, feed_start-3, col_count, 'Mass balances')

        self.fill_cell_bold(sheet, feed_start-2, col_count, 'Raw materials')
        for head in headers:
            self.fill_cell_bold(sheet, feed_start-1, col_count, head)
            col_count += 1

        # Print product headers 
        prod_start = feed_start + feed_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, prod_start-2, col_count, 'Products')
        for head in headers:
            self.fill_cell_bold(sheet, prod_start-1, col_count, head)
            col_count += 1

        # Print waste headers
        wast_start = prod_start + product_count + 3
        col_count = 2
        self.fill_cell_bold(sheet, wast_start-2, col_count, 'Waste streams')
        for head in headers:
            self.fill_cell_bold(sheet, wast_start-1, col_count, head)
            col_count += 1
        
        # Print Balance check
        check_start = wast_start + waste_count
        self.fill_cell_bold(sheet,check_start-7, 9,'Balance check')
        self.fill_cell_bold(sheet, check_start-6, 9,'Total inputs')
        self.fill_cell_bold(sheet, check_start-6, 10, 'Total F.rate ktonne/y ')
        self.fill_cell_bold(sheet, check_start-5, 9, 'Raw materials')
        sheet.cell(row=check_start-5, column= 10).value = feed_mass
        self.fill_cell_bold(sheet, check_start-4, 9, 'Products')
        sheet.cell(row=check_start-4, column= 10).value = product_mass
        self.fill_cell_bold(sheet, check_start-3, 9, 'Waste streams')
        sheet.cell(row=check_start-3, column= 10).value = waste_mass
        self.fill_cell_bold(sheet, check_start-2, 9, 'Balance check')
        sheet.cell(row=check_start-2, column= 10).value = feed_mass - product_mass - waste_mass
        self.fill_cell_bold(sheet, check_start-1, 9, 'Balance error %')
        sheet.cell(row=check_start-1, column= 10).value = (feed_mass - product_mass - waste_mass)/feed_mass*100


        wb.save(work_book)



    def Print_Energy(self, utilities, work_book):
        
        sheet_name = 'Energy balances v2' 
        
        wb = load_workbook(work_book)
        try:
            sheet = wb[sheet_name]
            wb.remove(sheet)
            sheet = wb.create_sheet(sheet_name)
        except:
            sheet = wb.create_sheet(sheet_name)

        start = 18
        util_count = 0
        headleft = ['Utility type','Input - TJ/y', 'Output TJ/y', 'Net TJ/y', 'Remark']
        headright = ['Utility type', 'Inlet temperature (C)', 'Outlet temperature (C)', 'Inlet pressure (bar)', 
        'Outlet pressure', 'Heat capacity/latent heat/lower heating', 'Unit','Remark']
        Utility = ['Low low pressure steam', 'Low pressure steam', 'Medium pressure steam', 'High pressure steam', 
        'Refrigerant1', 'Refrigerant2', 'Refrigerant3', 'Refrigerant4', 'Electricity', 'Natural gas','Cooling water']
        
        count_col = 2
        self.fill_cell_bold(sheet, start-2, count_col, 'Overall energy needs')
        for head in headleft:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1
        
        count_col = 10
        self.fill_cell_bold(sheet, start-2, count_col, 'Utility process conditions')
        for head in headright:
            self.fill_cell_bold(sheet, start-1, count_col, head)
            count_col += 1
        
        row_count = start
        for util in Utility:

            row_count += 1 
            util_count += 1 
                




        count_right = 0
        count_left = 0
        count_llps, count_lps, count_mps, count_hps = 0, 0, 0, 0
        count_frig1, count_frig2, count_frig3, count_frig4 = 0, 0, 0, 0
        count_elec, count_natgas, count_cw = 0, 0, 0
        rw = start + util_count + 6
        
        for util, block1 in utilities.items():
            for name, data in block1.items():
                if (util == 'LLPS' or util == 'LLPS-GEN'):
                    cur_row = rw + count_left
                    ii = 0
                    count_left += 1
                    count_llps += 1
                elif (util == 'LPS' or util == 'LPS-GEN'):
                    cur_row = rw + count_left + 3 
                    ii = 0
                    count_left += 1
                    count_lps += 1
                elif (util == 'MPS' or util == 'MPS-GEN'):
                    cur_row = rw + count_left + 6
                    ii = 0
                    count_left += 1
                    count_mps += 1
                elif (util == 'HPS' or util == 'HPS-GEN'):
                    cur_row = rw + count_left + 9
                    ii = 0
                    count_left += 1
                    count_hps += 1
                elif util == 'RF1':
                    cur_row = rw + count_left + 12
                    ii = 0
                    count_left += 1
                    count_frig1 += 1
                elif util == 'RF2':
                    cur_row = rw + count_left + 15
                    ii = 0
                    count_left += 1
                    count_frig2 += 1
                elif util == 'RF3':
                    cur_row = rw + count_left + 18
                    ii = 0
                    count_left += 1
                    count_frig3 += 1
                elif util == 'RF4':
                    cur_row = rw + count_left + 21
                    ii = 0
                    count_left += 1
                    count_frig4 += 1
                elif util == 'ELECTRIC':
                    cur_row = rw + count_right
                    ii = 8
                    count_right += 1
                    count_elec += 1
                elif util == 'NATGAS':
                    cur_row = rw + count_right + 3
                    ii = 8
                    count_right += 1
                    count_natgas += 1
                elif util == 'CW':
                    cur_row = rw + count_right + 6
                    ii = 8
                    count_right += 1
                    count_cw += 1
            
                if util == 'ELECTRIC':
                    sheet.cell(row=cur_row, column=2+ii).value = name
                    sheet.cell(row=cur_row, column=4+ii).value = -data.duty * 4186.8
                elif util =='NATGAS':
                    sheet.cell(row=cur_row, column=2+ii).value = name
                    sheet.cell(row=cur_row, column=6+ii).value = -data.usage * 8000 / 1000 / 1000

                else:
                    if '-GEN' in util:
                        sheet.cell(row=cur_row, column=2+ii).value = name
                        sheet.cell(row=cur_row, column=4+ii).value = data.duty * 4186.8
                        sheet.cell(row=cur_row, column=5+ii).value = data.duty * 4186.8 * 8000 * 1E-6
                        sheet.cell(row=cur_row, column=6+ii).value = data.usage * 8000 / 1000 / 1000
                    else:
                        sheet.cell(row=cur_row, column=2+ii).value = name
                        sheet.cell(row=cur_row, column=4+ii).value = -data.duty * 4186.8
                        sheet.cell(row=cur_row, column=5+ii).value = -data.duty * 4186.8 * 8000 * 1E-6
                        sheet.cell(row=cur_row, column=6+ii).value = -data.usage * 8000 / 1000 / 1000
    
        headers = ['Unit name','Unit description','Duty -MJ/hr', 'Duty -TJ/y', 'Mass -ktonne/y', 'Remark']
        headelec = ['Unit name','Unit description','Power -kW', 'Energy -GWh/y', 'Energy - TJ/y', 'Remark']

        # Print low pressure steam header
        count_col = 2
        lowlowpres_start = rw
        self.fill_cell_bold(sheet, lowlowpres_start-3, count_col, 'Breakdown of requirements')
        self.fill_cell_bold(sheet, lowlowpres_start-2, count_col, 'Low Low pressure steam')
        for head in headers:
            self.fill_cell_bold(sheet, lowlowpres_start-1, count_col, head)
            count_col += 1
        # Print low pressure steam header
        count_col = 2
        lowpres_start = lowlowpres_start + count_llps + 3
        self.fill_cell_bold(sheet, lowpres_start-2, count_col, 'Low pressure steam')
        for head in headers:
            self.fill_cell_bold(sheet, lowpres_start-1, count_col, head)
            count_col += 1
        # Print medium pressure steam header
        count_col = 2
        mp_start = lowpres_start + count_lps + 3
        self.fill_cell_bold(sheet, mp_start-2, count_col, 'Medium pressure steam')
        for head in headers:
            self.fill_cell_bold(sheet, mp_start-1, count_col, head)
            count_col += 1
        # Print high pressure steam header
        count_col = 2
        hp_start = mp_start + count_mps + 3 
        self.fill_cell_bold(sheet, hp_start-2, count_col, 'High pressure steam')
        for head in headers:
            self.fill_cell_bold(sheet, hp_start-1, count_col, head)
            count_col += 1
        # Print refrigerator 1 header
        count_col = 2
        frig_start1 = hp_start + count_hps + 3
        self.fill_cell_bold(sheet, frig_start1-2, count_col, 'Refrigerator 1')
        for head in headers:
            self.fill_cell_bold(sheet, frig_start1-1, count_col, head)
            count_col += 1
        # Print refrigerator 2 header
        count_col = 2
        frig_start2 = frig_start1 + count_frig1 + 3
        self.fill_cell_bold(sheet, frig_start2-2, count_col, 'Refrigerator 2')
        for head in headers:
            self.fill_cell_bold(sheet, frig_start2-1, count_col, head)
            count_col += 1
        # Print refrigerator 3 header
        count_col = 2
        frig_start3 = frig_start2 + count_frig2 + 3
        self.fill_cell_bold(sheet, frig_start3-2, count_col, 'Refrigerator 3')
        for head in headers:
            self.fill_cell_bold(sheet, frig_start3-1, count_col, head)
            count_col += 1
        # Print refrigerator 4 header
        count_col = 2
        frig_start4 = frig_start3 + count_frig3 + 3
        self.fill_cell_bold(sheet, frig_start4-2, count_col, 'Refrigerator 4')
        for head in headers:
            self.fill_cell_bold(sheet, frig_start4-1, count_col, head)
            count_col += 1

        # Print electricity header
        count_col = 10
        elec_start = rw
        self.fill_cell_bold(sheet, elec_start-2, count_col, 'Electricity')
        for head in headelec:
            self.fill_cell_bold(sheet, elec_start-1, count_col, head)
            count_col += 1
        # Print natural gas header
        count_col = 10
        natgas_start = elec_start+ count_elec + 3
        self.fill_cell_bold(sheet, natgas_start-2, count_col, 'Natural gas')
        for head in headers:
            self.fill_cell_bold(sheet, natgas_start-1, count_col, head)
            count_col += 1
        # Print cool water header
        count_col = 10
        water_start = natgas_start + count_natgas + 3
        self.fill_cell_bold(sheet, water_start-2, count_col, 'Cooling water')
        for head in headers:
            self.fill_cell_bold(sheet, water_start-1, count_col, head)
            count_col += 1

        wb.save(work_book)

    
    def fill_cell(self, sheet, row1, column1, value):
        sheet.cell(row=row1, column=column1)
    
    def fill_cell_bold(self, sheet, row1, column1, value):
        c = sheet.cell(row=row1, column=column1)
        c.value = value
        c.font = Font(bold=True)